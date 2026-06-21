"""
eeprom_to_excel.py
Baca EEPROM dump dari STM32 via Serial, parse struct LF robot, export ke Excel.

Requires:
    pip install pyserial openpyxl

Struct layout (per slot = 2259 bytes):
    0x0000 - 0x000E : GlobalConfig       (15 bytes)
    0x000F - 0x08A6 : CounterParam x100  (2200 bytes, 22 bytes each)
    0x08A7 - 0x08B4 : SensorThresh x14   (14 bytes)
    0x08B5 - 0x08D2 : CheckpointParam x10(30 bytes, 3 bytes each)

7 slot total, mulai dari address 0:
    Slot 0: 0x0000, Slot 1: 0x08D3, dst.
"""

import struct
import time
import serial
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── CONFIG ──────────────────────────────────────────────────────────────────
PORT        = "COM4"        # ganti sesuai port kamu
BAUDRATE    = 115200
TIMEOUT     = 10            # detik
OUTPUT_FILE = "eeprom_dump.xlsx"
OUTPUT_BIN  = "eeprom_dump.bin"

# ─── KONSTANTA STRUCT (ikut sizeof() dari C, #pragma pack(1)) ─────────────────
COUNTER_MAX       = 100
CP_MAX            = 10
SENSOR_COUNT      = 14
EEPROM_SLOT_COUNT = 14

# GlobalConfig: 15 bytes
# mode(u8) speed_mode(u8) kp(u8) kd(u8) line(u8) periode(u16) t_blank(u16)
# pwm_freq_khz(f32) mem_slot(u8) mirrored(u8/bool)
GLOBAL_CONFIG_FMT  = "<BBBBBHHfBB"
GLOBAL_CONFIG_SIZE = struct.calcsize(GLOBAL_CONFIG_FMT)   # = 15

# CounterParam: 22 bytes
# timer(u16) speed1(u8) speed2(u8) kp(u8) trigger(u16) decision(u8)
# delay_type(u8) delay_ms(u16) motor_l(i16) motor_r(i16)
# Encd_l(i16) Encd_r(i16) Encd_b(i16) Line_C(u8)
COUNTER_PARAM_FMT  = "<HBBBHBBHhhhhhB"
COUNTER_PARAM_SIZE = struct.calcsize(COUNTER_PARAM_FMT)   # = 22

# CheckpointParam: 3 bytes
# counter_pos(u8) timer_cp(u16)
CP_PARAM_FMT  = "<BH"
CP_PARAM_SIZE = struct.calcsize(CP_PARAM_FMT)             # = 3

# Offset per slot (sama persis dengan eeprom.h)
EEPROM_OFFSET_CONFIG  = 0
EEPROM_OFFSET_COUNTER = EEPROM_OFFSET_CONFIG  + GLOBAL_CONFIG_SIZE
EEPROM_OFFSET_SENSOR  = EEPROM_OFFSET_COUNTER + COUNTER_PARAM_SIZE * COUNTER_MAX
EEPROM_OFFSET_CP      = EEPROM_OFFSET_SENSOR  + SENSOR_COUNT
EEPROM_SLOT_SIZE      = EEPROM_OFFSET_CP      + CP_PARAM_SIZE * CP_MAX

# ─── ENUM MAPS (nilai asli dari types.h) ──────────────────────────────────────
ROBOT_MODE = {0: "NORMAL", 1: "COUNTER"}

# LINE_AUTO=0, LINE_HITAM=4, LINE_PUTIH=8
LINE_COLOR = {0: "AUTO", 4: "HITAM", 8: "PUTIH"}

DECISION = {
    0: "LOST", 1: "FREE", 2: "BELOK_KIRI",
    3: "BELOK_KANAN", 4: "STOP"
}

DELAY_TYPE = {0: "A", 1: "B"}

LINE_COUNTER = {
    0: "AUTO_NORMAL",  1: "AUTO_CENTER",  2: "AUTO_LEFT",   3: "AUTO_RIGHT",
    4: "BLACK_NORMAL", 5: "BLACK_CENTER", 6: "BLACK_LEFT",  7: "BLACK_RIGHT",
    8: "WHITE_NORMAL", 9: "WHITE_CENTER", 10: "WHITE_LEFT", 11: "WHITE_RIGHT",
}

# TRIGGER_TIMER = 0b11111111111111 = 16383
# TRIGGER_TICK  = 0b11111111111110 = 16382
# TRIGGER_BLANK = 0
TRIGGER_TIMER = 0b11111111111111
TRIGGER_TICK  = 0b11111111111110

_TRIG_VAL = [
    1<<9, 1<<10, 1<<11, 1<<12, 1<<13,          # L1-L5
    1<<4, 1<<3,  1<<2,  1<<1,  1<<0,            # R1-R5
    (1<<9)|(1<<4),  (1<<10)|(1<<3),             # LR1-LR5
    (1<<11)|(1<<2), (1<<12)|(1<<1), (1<<13)|(1<<0),
    0,              # BLANK
    TRIGGER_TIMER,  # TIMER
    TRIGGER_TICK,   # TICK/Encoder
]
_TRIG_NAME = [
    "L1","L2","L3","L4","L5",
    "R1","R2","R3","R4","R5",
    "LR1","LR2","LR3","LR4","LR5",
    "BLANK","TIMER","ENC",
]

def decode_trigger(val):
    for v, n in zip(_TRIG_VAL, _TRIG_NAME):
        if v == val:
            return n
    return f"SENSOR 0b{val:014b}"

# ─── PARSER ───────────────────────────────────────────────────────────────────

def parse_global_config(data, offset):
    raw = data[offset: offset + GLOBAL_CONFIG_SIZE]
    (mode, speed_mode, kp, kd, line, periode, t_blank,
     pwm_freq, mem_slot, mirrored) = struct.unpack_from(GLOBAL_CONFIG_FMT, raw)
    return {
        "mode":         ROBOT_MODE.get(mode, f"?{mode}"),
        "speed_mode":   speed_mode,
        "kp":           kp,
        "kd":           kd,
        "line":         LINE_COLOR.get(line, f"?{line}"),
        "periode":      periode,
        "t_blank":      t_blank,
        "pwm_freq_khz": round(pwm_freq, 3),
        "mem_slot":     mem_slot,
        "mirrored":     bool(mirrored),
    }

def parse_counter_params(data, offset):
    counters = []
    for i in range(COUNTER_MAX):
        o = offset + i * COUNTER_PARAM_SIZE
        (timer, speed1, speed2, kp, trigger, decision,
         delay_type, delay_ms, motor_l, motor_r,
         encd_l, encd_r, encd_b, line_c) = struct.unpack_from(COUNTER_PARAM_FMT, data, o)
        counters.append({
            "idx":        i,
            "timer":      timer,
            "speed1":     speed1,
            "speed2":     speed2,
            "kp":         kp,
            "trigger":    decode_trigger(trigger),
            "trigger_raw": f"0x{trigger:04X}",
            "decision":   DECISION.get(decision, f"?{decision}"),
            "delay_type": DELAY_TYPE.get(delay_type, f"?{delay_type}"),
            "delay_ms":   delay_ms,
            "motor_l":    motor_l,
            "motor_r":    motor_r,
            "encd_l":     encd_l,
            "encd_r":     encd_r,
            "encd_b":     encd_b,
            "line_c":     LINE_COUNTER.get(line_c, f"?{line_c}"),
        })
    return counters

def parse_sensor_thresh(data, offset):
    return list(data[offset: offset + SENSOR_COUNT])

def parse_checkpoints(data, offset):
    cps = []
    for i in range(CP_MAX):
        o = offset + i * CP_PARAM_SIZE
        (counter_pos, timer_cp) = struct.unpack_from(CP_PARAM_FMT, data, o)
        cps.append({"idx": i, "counter_pos": counter_pos, "timer_cp": timer_cp})
    return cps

def parse_slot(data, slot_idx):
    base = slot_idx * EEPROM_SLOT_SIZE
    return {
        "config":      parse_global_config(data,  base + EEPROM_OFFSET_CONFIG),
        "counters":    parse_counter_params(data,  base + EEPROM_OFFSET_COUNTER),
        "sensors":     parse_sensor_thresh(data,   base + EEPROM_OFFSET_SENSOR),
        "checkpoints": parse_checkpoints(data,     base + EEPROM_OFFSET_CP),
    }

# ─── EXCEL WRITER ─────────────────────────────────────────────────────────────

COLOR_HEADER  = "1F3864"
COLOR_SUBHEAD = "2E75B6"
COLOR_ALT     = "D6E4F0"
COLOR_WHITE   = "FFFFFF"

def make_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def cell_style(ws, row, col, value, bold=False, bg=None, center=False, font_color="000000"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=bold, color=font_color, size=10)
    c.border = make_border()
    c.alignment = Alignment(horizontal="center" if center else "left",
                             vertical="center", wrap_text=True)
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    return c

def write_header_row(ws, row, headers, bg=COLOR_HEADER):
    for col, h in enumerate(headers, 1):
        cell_style(ws, row, col, h, bold=True, bg=bg, center=True, font_color="FFFFFF")

def write_sheet_config(ws, slot_data, slot_idx):
    ws.title = f"Slot{slot_idx}_Config"

    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value = f"SLOT {slot_idx} — GlobalConfig"
    c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", start_color=COLOR_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")

    write_header_row(ws, 2, ["Parameter", "Value"], bg=COLOR_SUBHEAD)

    cfg = slot_data["config"]
    rows = [
        ("mode",         cfg["mode"]),
        ("speed_mode",   cfg["speed_mode"]),
        ("kp",           cfg["kp"]),
        ("kd",           cfg["kd"]),
        ("line",         cfg["line"]),
        ("periode",      cfg["periode"]),
        ("t_blank",      cfg["t_blank"]),
        ("pwm_freq_khz", cfg["pwm_freq_khz"]),
        ("mem_slot",     cfg["mem_slot"]),
        ("mirrored",     str(cfg["mirrored"])),
    ]
    for i, (k, v) in enumerate(rows):
        bg = COLOR_ALT if i % 2 == 0 else COLOR_WHITE
        cell_style(ws, i + 3, 1, k, bg=bg)
        cell_style(ws, i + 3, 2, v, bg=bg, center=True)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 18

def write_sheet_counter(ws, slot_data, slot_idx):
    ws.title = f"Slot{slot_idx}_Counter"

    headers = [
        "#", "timer", "speed1", "speed2", "kp",
        "trigger", "trigger_raw", "decision", "delay_type", "delay_ms",
        "motor_l", "motor_r", "encd_l", "encd_r", "encd_b", "line_c"
    ]

    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    c = ws["A1"]
    c.value = f"SLOT {slot_idx} — CounterParam (100 entries)"
    c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", start_color=COLOR_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")

    write_header_row(ws, 2, headers, bg=COLOR_SUBHEAD)

    for i, cp in enumerate(slot_data["counters"]):
        row = i + 3
        bg  = COLOR_ALT if i % 2 == 0 else COLOR_WHITE
        vals = [
            cp["idx"], cp["timer"], cp["speed1"], cp["speed2"], cp["kp"],
            cp["trigger"], cp["trigger_raw"], cp["decision"],
            cp["delay_type"], cp["delay_ms"],
            cp["motor_l"], cp["motor_r"], cp["encd_l"], cp["encd_r"], cp["encd_b"],
            cp["line_c"]
        ]
        for col, v in enumerate(vals, 1):
            center = (col != 6)
            cell_style(ws, row, col, v, bg=bg, center=center)

    col_widths = [4, 7, 8, 8, 5, 12, 12, 13, 11, 10, 9, 9, 8, 8, 8, 13]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def write_sheet_sensor(ws, slot_data, slot_idx):
    ws.title = f"Slot{slot_idx}_Sensor"

    ws.merge_cells(f"A1:{get_column_letter(SENSOR_COUNT)}1")
    c = ws["A1"]
    c.value = f"SLOT {slot_idx} — Sensor Threshold (14 sensor)"
    c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", start_color=COLOR_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")

    headers = [f"S{i}" for i in range(SENSOR_COUNT)]
    write_header_row(ws, 2, headers, bg=COLOR_SUBHEAD)

    for col, val in enumerate(slot_data["sensors"], 1):
        cell_style(ws, 3, col, val, bg=COLOR_ALT, center=True)

    for i in range(1, SENSOR_COUNT + 1):
        ws.column_dimensions[get_column_letter(i)].width = 8

def write_sheet_checkpoint(ws, slot_data, slot_idx):
    ws.title = f"Slot{slot_idx}_CP"

    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value = f"SLOT {slot_idx} — CheckpointParam (10 entries)"
    c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", start_color=COLOR_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")

    write_header_row(ws, 2, ["#", "counter_pos", "timer_cp"], bg=COLOR_SUBHEAD)

    for i, cp in enumerate(slot_data["checkpoints"]):
        row = i + 3
        bg  = COLOR_ALT if i % 2 == 0 else COLOR_WHITE
        # counter_pos 0xFF = tidak diset
        pos_display = "--" if cp["counter_pos"] == 0xFF else cp["counter_pos"]
        cell_style(ws, row, 1, cp["idx"],     bg=bg, center=True)
        cell_style(ws, row, 2, pos_display,   bg=bg, center=True)
        cell_style(ws, row, 3, cp["timer_cp"], bg=bg, center=True)

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12

def write_summary(ws, all_slots):
    ws.title = "Summary"

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "EEPROM Dump Summary — LF Robot"
    c.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", start_color=COLOR_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")

    headers = ["Slot", "Mode", "Speed", "KP/KD", "Mirror", "Active?"]
    write_header_row(ws, 2, headers, bg=COLOR_SUBHEAD)

    for i, slot in enumerate(all_slots):
        cfg = slot["config"]
        row = i + 3
        bg  = COLOR_ALT if i % 2 == 0 else COLOR_WHITE
        active = "✓" if cfg["mem_slot"] == i else ""
        cell_style(ws, row, 1, i,                              bg=bg, center=True)
        cell_style(ws, row, 2, cfg["mode"],                    bg=bg, center=True)
        cell_style(ws, row, 3, cfg["speed_mode"],              bg=bg, center=True)
        cell_style(ws, row, 4, f"{cfg['kp']}/{cfg['kd']}",    bg=bg, center=True)
        cell_style(ws, row, 5, "YES" if cfg["mirrored"] else "NO", bg=bg, center=True)
        cell_style(ws, row, 6, active, bg=bg, center=True, bold=(active == "✓"))

    for col, w in zip("ABCDEF", [6, 10, 8, 10, 8, 8]):
        ws.column_dimensions[col].width = w

# ─── SERIAL READ ──────────────────────────────────────────────────────────────

def read_eeprom_from_serial(port, baudrate):
    print(f"[*] Membuka port {port} @ {baudrate}...")
    ser = serial.Serial(port, baudrate, timeout=TIMEOUT)
    time.sleep(2)

    ser.write(b'P')
    resp = ser.readline().decode('utf-8', errors='ignore').strip()
    if resp != "OK":
        print(f"[!] Ping gagal, response: {repr(resp)}")
        ser.close()
        return None
    print("[*] Ping OK, memulai baca EEPROM...")

    ser.write(b'R')

    size_bytes = ser.read(2)
    if len(size_bytes) < 2:
        print("[!] Gagal baca size")
        ser.close()
        return None
    total_size = (size_bytes[0] << 8) | size_bytes[1]
    print(f"[*] Menerima {total_size} bytes...")

    data = bytearray()
    while len(data) < total_size:
        chunk = ser.read(min(512, total_size - len(data)))
        if not chunk:
            print(f"[!] Timeout, baru dapat {len(data)}/{total_size} bytes")
            break
        data.extend(chunk)
        print(f"\r    {len(data)}/{total_size} bytes ({100*len(data)//total_size}%)", end="")

    print()

    marker = ser.read(2)
    ser.close()

    if marker != b'\xDE\xAD':
        print(f"[!] End marker salah: {marker.hex()}, tapi data tetap diproses")

    return bytes(data)

# ─── BUILD WORKBOOK ───────────────────────────────────────────────────────────

def build_workbook(raw):
    min_size = EEPROM_SLOT_SIZE * EEPROM_SLOT_COUNT
    if len(raw) < min_size:
        print(f"[!] Data kurang: {len(raw)} bytes, minimal {min_size}")
        return None

    print(f"[*] Parsing {EEPROM_SLOT_COUNT} slot "
          f"(slot size={EEPROM_SLOT_SIZE}, counter size={COUNTER_PARAM_SIZE})...")

    all_slots = [parse_slot(raw, i) for i in range(EEPROM_SLOT_COUNT)]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    write_summary(wb.create_sheet(), all_slots)
    for i, slot in enumerate(all_slots):
        write_sheet_config(     wb.create_sheet(), slot, i)
        write_sheet_counter(    wb.create_sheet(), slot, i)
        write_sheet_sensor(     wb.create_sheet(), slot, i)
        write_sheet_checkpoint( wb.create_sheet(), slot, i)

    return wb

# ─── SAVE BIN (backup 1:1) ────────────────────────────────────────────────────

def save_bin(raw, filepath):
    """Simpan raw EEPROM bytes 1:1 sebagai file .bin untuk backup / flash ulang."""
    with open(filepath, "wb") as f:
        f.write(raw)
    print(f"[✓] Backup bin  : {filepath}  ({len(raw)} bytes)")


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    raw = read_eeprom_from_serial(PORT, BAUDRATE)
    if raw is None:
        return

    # Simpan backup .bin dulu sebelum parse — aman walau parse gagal
    save_bin(raw, OUTPUT_BIN)

    wb = build_workbook(raw)
    if wb is None:
        return

    wb.save(OUTPUT_FILE)
    print(f"[✓] Excel       : {OUTPUT_FILE}")
    print(f"    Total sheet : {len(wb.sheetnames)}")


# ─── WRITE EEPROM VIA SERIAL ─────────────────────────────────────────────────

def write_eeprom_to_serial(port, baudrate, bin_filepath):
    with open(bin_filepath, "rb") as f:
        raw = f.read()

    if len(raw) != 32768:
        print(f"[!] Ukuran file salah: {len(raw)} bytes, harus 32768")
        return

    print(f"[*] Membuka port {port} @ {baudrate}...")
    ser = serial.Serial(port, baudrate, timeout=10)
    time.sleep(2)

    # Ping
    ser.write(b'P')
    resp = ser.readline().decode('utf-8', errors='ignore').strip()
    if resp != "OK":
        print(f"[!] Ping gagal: {repr(resp)}")
        ser.close()
        return

    # Kirim command Write
    ser.write(b'W')

    # Tunggu ACK siap dari STM32
    ack = ser.read(1)
    if ack != b'K':
        print(f"[!] ACK salah: {ack}")
        ser.close()
        return
    print(f"[*] STM32 siap, mengirim {len(raw)} bytes...")

    # Kirim data per 64 byte, tunggu ACK 'K' per page dari ESP32
    chunk_size = 64
    total_pages = (len(raw) + chunk_size - 1) // chunk_size

    for page_idx, i in enumerate(range(0, len(raw), chunk_size)):
        chunk = raw[i : i + chunk_size]
        ser.write(chunk)
        ser.flush()

        # Tunggu ACK per page — ESP32 kirim 'K' setelah page write selesai
        ack = ser.read(1)
        if ack != b'K':
            print(f"\n[!] ACK page {page_idx} salah: {ack}")
            ser.close()
            return

        done = min(i + chunk_size, len(raw))
        print(f"\r    {done}/{len(raw)} bytes ({100 * done // len(raw)}%)", end="")

    print()

    # Tunggu konfirmasi done dari STM32
    result = ser.read(1)
    ser.close()

    if result == b'D':
        print("[✓] Write selesai!")
    else:
        print(f"[!] Respons tidak dikenal: {result}")


# ─── ENTRY POINT ──────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys

    if len(sys.argv) > 1 and sys.argv[1].endswith(".bin"):
        if "--write" in sys.argv:
            # Flash .bin ke EEPROM via serial
            # Usage: python script.py eeprom_dump.bin --write
            write_eeprom_to_serial(PORT, BAUDRATE, sys.argv[1])
        else:
            # Baca .bin → generate Excel saja (tidak overwrite .bin sumber)
            # Usage: python script.py eeprom_dump.bin
            print(f"[*] Mode file: membaca {sys.argv[1]}")
            with open(sys.argv[1], "rb") as f:
                raw = f.read()
            wb = build_workbook(raw)
            if wb:
                wb.save(OUTPUT_FILE)
                print(f"[✓] Excel       : {OUTPUT_FILE}")
    else:
        # Baca dari serial → simpan .bin + .xlsx
        # Usage: python script.py
        main()